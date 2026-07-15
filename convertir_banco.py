"""Regenera data/preguntas.json desde el Excel del banco DGESTTLA."""
from __future__ import annotations

import argparse
import json
import random
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Instala openpyxl: pip install openpyxl")


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def pick_distractors(correct: str, ambito: str, tipo: str, by_ambito, by_tipo, all_ans, n=3):
    correct_n = norm(correct)
    seen = {correct_n}
    picked = []
    for pool in (by_ambito[ambito], by_tipo[tipo], all_ans):
        candidates = [x for x in pool if norm(x) not in seen]
        candidates.sort(key=lambda x: abs(len(x) - len(correct)))
        random.shuffle(candidates)
        for c in candidates:
            if norm(c) in seen:
                continue
            picked.append(c)
            seen.add(norm(c))
            if len(picked) >= n:
                return picked
    while len(picked) < n:
        picked.append(f"Ninguna de las anteriores ({len(picked) + 1})")
    return picked[:n]


def convert(xlsx: Path, out: Path) -> int:
    random.seed(42)
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active

    rows = []
    for r in range(2, ws.max_row + 1):
        q = ws.cell(r, 2).value
        a = ws.cell(r, 3).value
        if not q or not a:
            continue
        oid = ws.cell(r, 1).value
        rows.append(
            {
                "id": int(oid) if oid is not None else r - 1,
                "pregunta": str(q).strip(),
                "respuesta": str(a).strip(),
                "fuente": str(ws.cell(r, 4).value or "").strip(),
                "tipo": str(ws.cell(r, 5).value or "").strip(),
                "ambito": str(ws.cell(r, 6).value or "").strip(),
            }
        )

    by_ambito = defaultdict(list)
    by_tipo = defaultdict(list)
    all_ans = []
    for row in rows:
        by_ambito[row["ambito"]].append(row["respuesta"])
        by_tipo[row["tipo"]].append(row["respuesta"])
        all_ans.append(row["respuesta"])

    preguntas = []
    for row in rows:
        distractors = pick_distractors(
            row["respuesta"], row["ambito"], row["tipo"], by_ambito, by_tipo, all_ans
        )
        opciones = [row["respuesta"], *distractors]
        rng = random.Random(row["id"] * 9973 + 17)
        order = list(range(4))
        rng.shuffle(order)
        ordered = [opciones[i] for i in order]
        correct_letter = chr(ord("A") + ordered.index(row["respuesta"]))
        preguntas.append(
            {
                "id": row["id"],
                "pregunta": row["pregunta"],
                "opciones": {
                    "A": ordered[0],
                    "B": ordered[1],
                    "C": ordered[2],
                    "D": ordered[3],
                },
                "correcta": correct_letter,
                "fuente": row["fuente"],
                "tipo": row["tipo"],
                "ambito": row["ambito"],
            }
        )

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        json.dumps({"total": len(preguntas), "preguntas": preguntas}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return len(preguntas)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "xlsx",
        nargs="?",
        default=str(Path.home() / "Downloads" / "Banco_Definitivo_DGESTTLA_2026_708_Preguntas.xlsx"),
    )
    parser.add_argument("-o", "--out", default=str(Path(__file__).parent / "data" / "preguntas.json"))
    args = parser.parse_args()
    n = convert(Path(args.xlsx), Path(args.out))
    print(f"Generadas {n} preguntas en {args.out}")


if __name__ == "__main__":
    main()
